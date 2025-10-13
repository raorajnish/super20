from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import Enquiry, Admission, Faculty, Lecture, AttendanceRecord, Payment
from .forms import EnquiryForm, AdmissionForm, EnquiryUpdateForm
from django.utils import timezone
from django.contrib.auth.models import User

def home(request):
    """Home page with hero section and navigation"""
    return render(request, 'admissions/home.html')

def enquiry_form(request):
    """Enquiry form submission"""
    if request.method == 'POST':
        form = EnquiryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Enquiry submitted successfully! We will contact you soon.')
            return redirect('home')
    else:
        form = EnquiryForm()
    
    return render(request, 'admissions/enquiry_form.html', {'form': form})

def admission_form(request):
    """Admission form submission"""
    if request.method == 'POST':
        form = AdmissionForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Admission form submitted successfully! We will review and contact you soon.')
            return redirect('home')
    else:
        form = AdmissionForm()
    
    return render(request, 'admissions/admission_form.html', {'form': form})

def admin_login(request):
    """Admin login view"""
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None and user.is_staff:
                login(request, user)
                messages.success(request, f'Welcome back, {username}!')
                return redirect('dashboard')
            else:
                messages.error(request, 'Invalid credentials or insufficient permissions.')
    else:
        form = AuthenticationForm()
    
    return render(request, 'admissions/admin_login.html', {'form': form})

@login_required
def admin_logout(request):
    """Admin logout view"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('home')

@login_required
def dashboard(request):
    """Admin dashboard with statistics"""
    if not request.user.is_staff:
        return redirect('home')
    total_enquiries = Enquiry.objects.count()
    total_admissions = Admission.objects.count()
    converted_enquiries = Enquiry.objects.filter(status='converted').count()
    conversion_rate = (converted_enquiries / total_enquiries * 100) if total_enquiries > 0 else 0
    
    # Recent activities
    recent_enquiries = Enquiry.objects.order_by('-enquiry_date')[:5]
    recent_admissions = Admission.objects.order_by('-submitted_at')[:5]
    
    # Handle quick-create faculty
    if request.method == 'POST' and request.POST.get('action') == 'create_faculty':
        fname = request.POST.get('full_name', '').strip()
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        phone = request.POST.get('phone_number', '').strip()
        if not (fname and username and password):
            messages.error(request, 'Full name, username and password are required.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists. Choose a different one.')
        else:
            user = User.objects.create_user(username=username, password=password, is_active=True)
            fac = Faculty.objects.create(user=user, full_name=fname, phone_number=phone, is_active=True)
            messages.success(request, f'Faculty "{fac.full_name}" created with username "{username}".')
            return redirect('dashboard')

    faculties = Faculty.objects.select_related('user').order_by('-id')[:10]
    # Prepare current month payment snapshots for quick view
    from datetime import date
    month_start = date.today().replace(day=1)
    faculty_cards = []
    for f in faculties:
        payment, _ = Payment.objects.get_or_create(faculty=f, month=month_start, defaults={'per_lecture_rate': 0, 'amount_paid': 0})
        faculty_cards.append({
            'obj': f,
            'lectures_count': payment.lectures_count,
            'per_lecture_rate': payment.per_lecture_rate,
            'amount_due': payment.amount_due,
            'amount_paid': payment.amount_paid,
            'balance': payment.balance,
        })

    context = {
        'total_enquiries': total_enquiries,
        'total_admissions': total_admissions,
        'converted_enquiries': converted_enquiries,
        'conversion_rate': round(conversion_rate, 1),
        'recent_enquiries': recent_enquiries,
        'recent_admissions': recent_admissions,
        'faculties': faculties,
        'faculty_cards': faculty_cards,
    }
    return render(request, 'admissions/admin_dashboard.html', context)

@login_required
def enquiry_list(request):
    """Enquiry management page with filters and search"""
    enquiries = Enquiry.objects.all().order_by('-enquiry_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        enquiries = enquiries.filter(
            Q(student_name__icontains=search_query) |
            Q(guardian_name__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        enquiries = enquiries.filter(status=status_filter)
    
    # Date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        enquiries = enquiries.filter(enquiry_date__date__gte=date_from)
    if date_to:
        enquiries = enquiries.filter(enquiry_date__date__lte=date_to)
    
    # Pagination
    paginator = Paginator(enquiries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'status_choices': Enquiry.STATUS_CHOICES,
    }
    return render(request, 'admissions/enquiry_list.html', context)

@login_required
def edit_enquiry(request, id):
    """Edit enquiry notes and status"""
    enquiry = get_object_or_404(Enquiry, id=id)
    
    if request.method == 'POST':
        form = EnquiryUpdateForm(request.POST, instance=enquiry)
        if form.is_valid():
            form.save()
            messages.success(request, 'Enquiry updated successfully!')
            return redirect('enquiry_list')
    else:
        form = EnquiryUpdateForm(instance=enquiry)
    
    context = {
        'form': form,
        'enquiry': enquiry,
    }
    return render(request, 'admissions/edit_enquiry.html', context)

@login_required
def admission_list(request):
    """Admission management page with filters and search"""
    admissions = Admission.objects.all().order_by('-submitted_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        admissions = admissions.filter(
            Q(name__icontains=search_query) |
            Q(surname__icontains=search_query) |
            Q(mobile_1__icontains=search_query) |
            Q(school_college__icontains=search_query)
        )
    
    # Standard filter
    standard_filter = request.GET.get('standard', '')
    if standard_filter:
        admissions = admissions.filter(standard=standard_filter)
    
    # Batch filter
    batch_filter = request.GET.get('batch', '')
    if batch_filter:
        admissions = admissions.filter(batch__icontains=batch_filter)
    
    # Date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        admissions = admissions.filter(submitted_at__date__gte=date_from)
    if date_to:
        admissions = admissions.filter(submitted_at__date__lte=date_to)
    
    # Pagination
    paginator = Paginator(admissions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'standard_filter': standard_filter,
        'batch_filter': batch_filter,
        'date_from': date_from,
        'date_to': date_to,
        'standard_choices': Admission.STANDARD_CHOICES,
    }
    return render(request, 'admissions/admission_list.html', context)

@login_required
def admission_detail(request, id):
    """Student profile page with full details"""
    admission = get_object_or_404(Admission, id=id)
    return render(request, 'admissions/admission_detail.html', {'admission': admission})

def about_us(request):
    """About us page"""
    return render(request, 'admissions/about_us.html')

def contact(request):
    """Contact page"""
    return render(request, 'admissions/contact.html')

@login_required
def export_enquiries(request):
    """Export enquiries to Excel file organized by preferred course"""
    # Get all enquiries
    enquiries = Enquiry.objects.all().order_by('preferred_course', 'enquiry_date')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Enquiries Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Super20 Academy - Enquiries Report (Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')})"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add headers
    headers = [
        'ID', 'Student Name', 'Guardian Name', 'Phone Number', 
        'Preferred Course', 'Enquiry Date', 'Status', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    row = 4
    for enquiry in enquiries:
        ws.cell(row=row, column=1, value=enquiry.id).border = thin_border
        ws.cell(row=row, column=2, value=enquiry.student_name).border = thin_border
        ws.cell(row=row, column=3, value=enquiry.guardian_name).border = thin_border
        ws.cell(row=row, column=4, value=enquiry.phone_number).border = thin_border
        ws.cell(row=row, column=5, value=enquiry.get_preferred_course_display()).border = thin_border
        ws.cell(row=row, column=6, value=enquiry.enquiry_date.strftime('%d/%m/%Y %H:%M')).border = thin_border
        ws.cell(row=row, column=7, value=enquiry.get_status_display()).border = thin_border
        ws.cell(row=row, column=8, value=enquiry.notes or '').border = thin_border
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Super20_Enquiries_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def export_admissions(request):
    """Export admissions to Excel file organized by standard"""
    # Get all admissions
    admissions = Admission.objects.all().order_by('standard', 'name')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Admissions Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f"Super20 Academy - Admissions Report (Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')})"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add headers
    headers = [
        'ID', 'Full Name', 'Standard', 'Batch', 'Stream', 'Date of Birth', 'Age',
        'Father Name', 'Mother Name', 'Mobile 1', 'Mobile 2', 'School/College'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    row = 4
    for admission in admissions:
        # Calculate age
        from datetime import date
        today = date.today()
        age = today.year - admission.date_of_birth.year - ((today.month, today.day) < (admission.date_of_birth.month, admission.date_of_birth.day))
        
        ws.cell(row=row, column=1, value=admission.id).border = thin_border
        ws.cell(row=row, column=2, value=admission.full_name()).border = thin_border
        ws.cell(row=row, column=3, value=admission.get_standard_display()).border = thin_border
        ws.cell(row=row, column=4, value=admission.batch).border = thin_border
        ws.cell(row=row, column=5, value=admission.get_stream_display() if admission.stream else '').border = thin_border
        ws.cell(row=row, column=6, value=admission.date_of_birth.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(row=row, column=7, value=age).border = thin_border
        ws.cell(row=row, column=8, value=admission.father_name).border = thin_border
        ws.cell(row=row, column=9, value=admission.mother_name).border = thin_border
        ws.cell(row=row, column=10, value=admission.mobile_1).border = thin_border
        ws.cell(row=row, column=11, value=admission.mobile_2 or '').border = thin_border
        ws.cell(row=row, column=12, value=admission.school_college).border = thin_border
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Super20_Admissions_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response


# -------------------- FACULTY AUTH AND DASHBOARD --------------------
def faculty_login(request):
    """Faculty login using credentials created by admin (User + Faculty)."""
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None and hasattr(user, 'faculty_profile') and user.is_active:
                login(request, user)
                messages.success(request, f'Welcome, {user.faculty_profile.full_name}!')
                return redirect('faculty_dashboard')
            messages.error(request, 'Invalid faculty credentials or account inactive.')
    else:
        form = AuthenticationForm()
    return render(request, 'admissions/faculty_login.html', {'form': form})


@login_required
def faculty_logout(request):
    if hasattr(request.user, 'faculty_profile') or request.user.is_staff:
        logout(request)
        messages.success(request, 'Logged out successfully.')
    return redirect('home')


@login_required
def faculty_dashboard(request):
    """Faculty dashboard showing upcoming and past lectures."""
    if not hasattr(request.user, 'faculty_profile'):
        return redirect('home')
    faculty = request.user.faculty_profile
    today = timezone.localdate()
    upcoming = Lecture.objects.filter(faculty=faculty, date__gte=today).order_by('date', 'start_time')
    past = Lecture.objects.filter(faculty=faculty, date__lt=today).order_by('-date', '-start_time')[:10]
    # Salary info for current month
    from datetime import date
    month_start = date.today().replace(day=1)
    payment, _ = Payment.objects.get_or_create(faculty=faculty, month=month_start, defaults={'per_lecture_rate': 0, 'amount_paid': 0})
    return render(request, 'admissions/faculty_dashboard.html', {
        'upcoming_lectures': upcoming,
        'recent_lectures': past,
        'payment': payment,
    })


@login_required
def faculty_profile(request, faculty_id):
    """Admin view: per-faculty profile, monthly payment management."""
    if not request.user.is_staff:
        return redirect('home')
    faculty = get_object_or_404(Faculty, id=faculty_id)
    from datetime import date
    month_start = date.today().replace(day=1)
    payment, _ = Payment.objects.get_or_create(
        faculty=faculty, month=month_start, defaults={'per_lecture_rate': 0, 'amount_paid': 0}
    )

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_rate':
            try:
                from decimal import Decimal, InvalidOperation
                rate_raw = request.POST.get('per_lecture_rate', '0').strip()
                rate_clean = rate_raw.replace(',', '')
                rate_val = Decimal(rate_clean)
                if rate_val < 0:
                    raise InvalidOperation
                payment.per_lecture_rate = rate_val
                payment.save()
                messages.success(request, 'Per-lecture rate updated.')
            except Exception:
                messages.error(request, 'Invalid rate value.')
        elif action == 'record_payment':
            try:
                from decimal import Decimal, InvalidOperation
                amount_raw = request.POST.get('amount', '0').strip()
                amount_clean = amount_raw.replace(',', '')
                amount_val = Decimal(amount_clean)
                if amount_val < 0:
                    raise InvalidOperation
                payment.amount_paid = (payment.amount_paid or 0) + amount_val
                payment.save()
                messages.success(request, 'Payment recorded.')
            except Exception:
                messages.error(request, 'Invalid payment amount.')
        return redirect('faculty_profile', faculty_id=faculty.id)

    # History: last 6 months
    history = Payment.objects.filter(faculty=faculty).order_by('-month')[:6]

    return render(request, 'admissions/faculty_profile.html', {
        'faculty': faculty,
        'payment': payment,
        'history': history,
    })


# -------------------- LECTURES CRUD AND ATTENDANCE --------------------
@login_required
def lecture_list(request):
    """Admin list of all lectures with filters. Faculty sees own lectures."""
    if hasattr(request.user, 'faculty_profile') and not request.user.is_staff:
        lectures = Lecture.objects.filter(faculty=request.user.faculty_profile)
    else:
        lectures = Lecture.objects.all()
    q = request.GET.get('q', '')
    if q:
        lectures = lectures.filter(Q(title__icontains=q) | Q(description__icontains=q) | Q(batch__icontains=q))
    paginator = Paginator(lectures.order_by('-date', '-start_time'), 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'admissions/lecture_list.html', {'page_obj': page_obj, 'q': q})


@login_required
def lecture_create(request):
    """Create a lecture (admin only)."""
    if not request.user.is_staff:
        return redirect('lecture_list')
    if request.method == 'POST':
        # Minimal inline form handling to avoid creating a new forms.py class for now
        data = request.POST
        faculty_id = data.get('faculty')
        faculty = get_object_or_404(Faculty, id=faculty_id)
        Lecture.objects.create(
            title=data.get('title'),
            description=data.get('description'),
            date=data.get('date'),
            start_time=data.get('start_time'),
            end_time=data.get('end_time'),
            standard=data.get('standard'),
            batch=data.get('batch'),
            faculty=faculty,
        )
        messages.success(request, 'Lecture created successfully.')
        return redirect('lecture_list')
    return render(request, 'admissions/lecture_form.html', {
        'faculties': Faculty.objects.filter(is_active=True).order_by('full_name'),
        'standard_choices': Admission.STANDARD_CHOICES,
    })


@login_required
def lecture_edit(request, lecture_id):
    if not request.user.is_staff:
        return redirect('lecture_list')
    lecture = get_object_or_404(Lecture, id=lecture_id)
    if request.method == 'POST':
        data = request.POST
        lecture.title = data.get('title')
        lecture.description = data.get('description')
        lecture.date = data.get('date')
        lecture.start_time = data.get('start_time')
        lecture.end_time = data.get('end_time')
        lecture.standard = data.get('standard')
        lecture.batch = data.get('batch')
        lecture.faculty = get_object_or_404(Faculty, id=data.get('faculty'))
        lecture.save()
        messages.success(request, 'Lecture updated successfully.')
        return redirect('lecture_detail', lecture_id=lecture.id)
    return render(request, 'admissions/lecture_form.html', {
        'lecture': lecture,
        'faculties': Faculty.objects.filter(is_active=True).order_by('full_name'),
        'standard_choices': Admission.STANDARD_CHOICES,
    })


@login_required
def lecture_delete(request, lecture_id):
    if not request.user.is_staff:
        return redirect('lecture_list')
    lecture = get_object_or_404(Lecture, id=lecture_id)
    lecture.delete()
    messages.success(request, 'Lecture deleted successfully.')
    return redirect('lecture_list')


@login_required
def lecture_detail(request, lecture_id):
    lecture = get_object_or_404(Lecture, id=lecture_id)
    # Authorization: staff or assigned faculty only
    if hasattr(request.user, 'faculty_profile') and lecture.faculty != request.user.faculty_profile and not request.user.is_staff:
        return redirect('faculty_dashboard')
    students = list(lecture.get_target_students_queryset())
    records = {r.student_id: r for r in AttendanceRecord.objects.filter(lecture=lecture)}
    for s in students:
        setattr(s, 'attendance_record', records.get(s.id))
    return render(request, 'admissions/lecture_detail.html', {
        'lecture': lecture,
        'students': students,
    })


@login_required
def lecture_attendance(request, lecture_id):
    lecture = get_object_or_404(Lecture, id=lecture_id)
    # Authorization
    if hasattr(request.user, 'faculty_profile') and lecture.faculty != request.user.faculty_profile and not request.user.is_staff:
        return redirect('faculty_dashboard')
    if request.method == 'POST':
        # Expect POST as dict of student_<id>=present/absent
        marked_by_faculty = request.user.faculty_profile if hasattr(request.user, 'faculty_profile') else None
        absentees = []
        for student in lecture.get_target_students_queryset():
            key = f'student_{student.id}'
            status = request.POST.get(key, 'present')
            record, _ = AttendanceRecord.objects.update_or_create(
                lecture=lecture,
                student=student,
                defaults={'status': status, 'marked_by': marked_by_faculty}
            )
            if status == 'absent':
                absentees.append(student)
        # Build WhatsApp-ready message
        date_str = lecture.date.strftime('%d-%m-%Y')
        title = lecture.title
        batch = lecture.batch
        standard = dict(Admission.STANDARD_CHOICES).get(lecture.standard, lecture.standard)
        if absentees:
            names = ', '.join([s.full_name() for s in absentees])
            whatsapp_text = f"Absentees for {title} ({standard}-{batch}) on {date_str}: {names}."
        else:
            whatsapp_text = f"All students present for {title} ({standard}-{batch}) on {date_str}."
        messages.success(request, 'Attendance submitted successfully. You can copy the WhatsApp message below.')
        return render(request, 'admissions/attendance_submitted.html', {
            'lecture': lecture,
            'whatsapp_text': whatsapp_text,
        })
    # GET -> show form
    students = list(lecture.get_target_students_queryset())
    records = {r.student_id: r for r in AttendanceRecord.objects.filter(lecture=lecture)}
    for s in students:
        setattr(s, 'attendance_record', records.get(s.id))
    return render(request, 'admissions/lecture_attendance.html', {
        'lecture': lecture,
        'students': students,
    })
